#!/usr/bin/env python3
"""
Builds a "SAP demand forecast" workbook: which companies are hiring for which SAP
solution/implementation/upgrade, what kind of project it is, and over what time frame.

IMPORTANT — method and limits (also written into the "Read me" sheet):
This is NOT a statistical/predictive forecast. It is a structured read-out of the same 65
live job postings gathered in `generate_xlsx_report.py`, re-classified by:
  - SAP solution detail (as stated in the posting) + a canonical solution category (so the
    summary sheet can actually aggregate, rather than 65 near-unique free-text labels)
  - a canonical project-type category (Implementation / Migration-Upgrade / Run & Support /
    Architecture & Strategy / Advisory / Security & Compliance / Project Mgmt & Training)
  - a time frame: an EXPLICIT date/duration when the posting stated one (3 of 65 roles do),
    else "Active now / ongoing" — meaning the company is hiring for this today, which is
    the only demand signal a job ad gives; it is not a projection of when the underlying
    project starts, ends, or how long it runs.
Treat the summary sheets as "what's being hired for this week across these 65 roles," not
a market-wide prediction — the sample is the curated Aug 2026 BE/NL SAP role list, not a
survey of the whole market. Categorization judgment calls (e.g. a role mentioning both SD
and MM) are the author's, not the posting's own words — cross-check the detail column or
the source posting if precision matters for a decision.
"""

import sys
from collections import defaultdict
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPORT_DATE = "2026-08-06"
root_dir = Path(__file__).parent.parent
findings_dir = root_dir / "findings"
findings_dir.mkdir(exist_ok=True, parents=True)
XLSX_OUTPUT_PATH = findings_dir / f"sap_forecast_be_nl_{REPORT_DATE}.xlsx"

# Canonical solution categories (assigned by primary module/technology focus of the role)
SOL_S4 = "S/4HANA Core / Transformation"
SOL_BTP = "SAP BTP & Integration"
SOL_DEV = "SAP ABAP / Custom Development"
SOL_SD = "SAP SD / CX (Sales & Distribution)"
SOL_MM = "SAP MM / Procurement (incl. Ariba)"
SOL_PP = "SAP PP / Manufacturing / MES"
SOL_WM = "SAP EWM / WM / TM (Warehouse & Transport)"
SOL_FI = "SAP FICO / Finance / Treasury"
SOL_EAM = "SAP EAM / PM (Asset Management)"
SOL_GTS = "SAP GTS (Global Trade Services)"
SOL_SEC = "SAP Security / GRC"
SOL_DATA = "SAP Analytics / BW / Group Reporting / Archiving"
SOL_B1 = "SAP Business One"
SOL_BASIS = "SAP Basis / NetWeaver Infrastructure"
SOL_AI = "AI + SAP / Cross-module Advisory"

# Canonical project-type categories
PT_IMPL = "Implementation"
PT_MIG = "Migration / Upgrade"
PT_RUN = "Run & Support"
PT_ARCH = "Architecture & Strategy"
PT_ADV = "Advisory"
PT_SEC = "Security & Compliance"
PT_PM = "Project Management & Training"

# (Role, Company, Country, Location, Solution detail (as stated in posting),
#  Solution category, Project type category, Time frame)
FORECAST = [
    ("SAP Integration Specialist", "Alliander", "NL", "Arnhem", "SAP BTP Integration Suite", SOL_BTP, PT_RUN, "Active now / ongoing"),
    ("SAP S/4HANA Roadmap & Business Transformation Advisor", "Accenture NL", "NL", "Amsterdam", "S/4HANA roadmap & strategy", SOL_S4, PT_ARCH, "Active now / ongoing"),
    ("SAP Solution Architect", "Brabant Water N.V.", "NL", "'s-Hertogenbosch", "SAP EAM & SCM architecture", SOL_EAM, PT_ARCH, "Active now / ongoing"),
    ("SAP-ontwikkelaar", "myBrand | Conclusion", "NL", "Maarssen", "SAP Cloud ERP (ABAP Cloud)", SOL_DEV, PT_MIG, "Active now / ongoing"),
    ("SAP/BTP Development Specialist", "Nyrstar", "NL", "Budel", "SAP BTP + custom ABAP (Finance/Metal Accounting)", SOL_BTP, PT_RUN, "Active now / ongoing"),
    ("SAP CAP Developer", "delaware the Netherlands", "NL", "Den Bosch", "SAP BTP (CAP/Fiori)", SOL_BTP, PT_IMPL, "Active now / ongoing"),
    ("SAP S/4HANA Data Transformation Advisor", "Accenture NL", "NL", "Amsterdam", "S/4HANA data/MDG/BW4HANA", SOL_DATA, PT_MIG, "Active now / ongoing"),
    ("Manager SAP S/4HANA Sales and Distribution", "Deloitte", "NL", "Amsterdam", "S/4HANA SD", SOL_SD, PT_IMPL, "Active now / ongoing"),
    ("SAP Production Consultant", "NTT DATA Business Solutions", "NL", "'s-Hertogenbosch", "SAP PP / PP-DS / APO / IBP", SOL_PP, PT_IMPL, "Active now / ongoing"),
    ("SAP Integration Consultant", "Accenture NL", "NL", "Amsterdam", "SAP CPI / PI-PO / API Management", SOL_BTP, PT_IMPL, "Active now / ongoing"),
    ("SAP Sales Distribution Consultant (SD/OTC)", "Accenture NL", "NL", "Amsterdam", "SAP SD / CX", SOL_SD, PT_IMPL, "Active now / ongoing"),
    ("SAP Development Architect & Team Lead", "Accenture NL", "NL", "Amsterdam", "S/4HANA + BTP (Clean Core)", SOL_BTP, PT_ARCH, "Active now / ongoing"),
    ("SAP Architect – NL", "Capgemini", "NL", "Utrecht", "S/4HANA Cloud architecture", SOL_S4, PT_ARCH, "Active now / ongoing"),
    ("SAP Manufacturing Manager", "Accenture NL", "NL", "Amsterdam", "S/4HANA Manufacturing / PP", SOL_PP, PT_IMPL, "Active now / ongoing"),
    ("SAP S/4 Digital EAM Consultant", "Accenture NL", "NL", "Amsterdam", "S/4HANA EAM", SOL_EAM, PT_IMPL, "Active now / ongoing"),
    ("SAP Security Lead", "Scandinavian Tobacco Group", "NL", "Waalre", "S/4HANA Security / GRC", SOL_SEC, PT_SEC, "Active now / ongoing"),
    ("Super User SAP", "Scania Production & Logistics NL", "NL", "Hasselt (Overijssel)", "SAP EWM in S/4HANA", SOL_WM, PT_RUN, "Active now / ongoing"),
    ("SAP Group Reporting & Public Cloud", "RED Global (client undisclosed)", "NL", "South Holland", "SAP Group Reporting (Consolidation)", SOL_DATA, PT_IMPL, "Immediate start (contract)"),
    ("SAP developer", "Stedin", "NL", "Delft", "SAP ABAP / Fiori (grid systems)", SOL_DEV, PT_RUN, "Active now / ongoing"),
    ("Technical Application manager", "Panda International (client undisclosed)", "NL", "Bilthoven", "SAP ECC / S/4HANA application mgmt", SOL_S4, PT_RUN, "Active now / ongoing"),
    ("Consultant SAP EAM", "Deloitte", "NL", "Amsterdam", "SAP EAM", SOL_EAM, PT_IMPL, "Active now / ongoing"),
    ("SAP Sourcing & Procurement Consultant", "Accenture NL", "NL", "Amsterdam", "S/4HANA + Ariba (Source-to-Pay)", SOL_MM, PT_IMPL, "Active now / ongoing"),
    ("SAP ABAP Developer", "Stedin", "NL", "Delft", "SAP ABAP / RAP / Fiori / CPI", SOL_DEV, PT_RUN, "Active now / ongoing"),
    ("SAP C4C consultant", "Stedin", "NL", "Rotterdam", "SAP C4C (CRM)", SOL_SD, PT_RUN, "Active now / ongoing"),
    ("AI Tech Consultant (ERP/SAP)", "BCG Platinion", "NL", "Amsterdam", "SAP (general) + AI/GenAI", SOL_AI, PT_ADV, "Active now / ongoing"),
    ("Lead AI Tech Architect (ERP/SAP)", "BCG Platinion", "NL", "Amsterdam", "S/4HANA + AI", SOL_AI, PT_ARCH, "Active now / ongoing"),
    ("Senior AI Tech Architect (ERP/SAP)", "BCG Platinion", "NL", "Amsterdam", "S/4HANA + AI", SOL_AI, PT_ARCH, "Active now / ongoing"),
    ("Senior SAP Developer", "vidaXL", "NL", "Venlo", "S/4HANA (logistics/supply chain apps)", SOL_DEV, PT_RUN, "Active now / ongoing"),
    ("SAP Developer", "Applied Medical", "NL", "Amersfoort", "SAP ABAP / Fiori (multi-module)", SOL_DEV, PT_RUN, "Active now / ongoing"),
    ("SAP Treasury Risk Management", "Accenture NL", "NL", "Amsterdam", "S/4HANA Treasury", SOL_FI, PT_IMPL, "Active now / ongoing"),
    ("SAP Analyst", "ThoughtLabs Belgium (client undisclosed)", "BE", "Brussels", "SAP MM + Ariba", SOL_MM, PT_RUN, "Active now / ongoing"),
    ("SAP PP/QM", "JoBBsquare België (client undisclosed)", "BE", "Antwerp", "SAP PP / QM", SOL_PP, PT_IMPL, "Active now / ongoing"),
    ("SAP Functioneel Analyst", "Madison Recruitment (client undisclosed)", "BE", "Aalst", "SAP SD", SOL_SD, PT_IMPL, "Active now / ongoing"),
    ("SAP GTS Consultant", "Deloitte Belgium", "BE", "Zaventem", "SAP GTS", SOL_GTS, PT_IMPL, "Active now / ongoing"),
    ("SAP TM Consultant", "Deloitte Belgium", "BE", "Zaventem", "SAP TM", SOL_WM, PT_IMPL, "Active now / ongoing"),
    ("System Exploitation Engineer – SAP BC Member", "Sibelga", "BE", "Brussels", "SAP NetWeaver / Basis / S/4HANA infra", SOL_BASIS, PT_RUN, "Active now / ongoing"),
    ("Functioneel SAP-analist", "Aures (client undisclosed)", "BE", "Wingene", "SAP (production/planning/quality)", SOL_PP, PT_IMPL, "Active now / ongoing"),
    ("SAP Enterprise Architect", "Robert Half (client undisclosed)", "BE", "Ghent", "S/4HANA", SOL_S4, PT_ARCH, "Multi-year engagement (~4 years, stated in posting)"),
    ("SAP BDC Lead", "KPMG Belgium", "BE", "Zaventem", "SAP Business Data Cloud / BTP", SOL_DATA, PT_ARCH, "Active now / ongoing"),
    ("Senior SAP Sales & Distribution Consultant", "delaware BeLux", "BE", "Ghent", "SAP SD", SOL_SD, PT_IMPL, "Active now / ongoing"),
    ("Responsable Fonctionnel SAP Quality / WMS", "Safran", "BE", "Liège", "S/4HANA QM / eWM", SOL_WM, PT_IMPL, "Active implementation, ongoing"),
    ("Responsable Fonctionnel SAP Finance", "Safran", "BE", "Liège", "S/4HANA FI-CO", SOL_FI, PT_IMPL, "Active implementation, ongoing"),
    ("SAP Architect (SD/MM)", "Amon (Lotus Bakeries)", "BE", "Kaprijke", "SAP SD / MM", SOL_SD, PT_ARCH, "Active now / ongoing"),
    ("SAP FICO Consultant", "Deloitte Belgium", "BE", "Zaventem", "SAP FICO (Tax Technology)", SOL_FI, PT_ADV, "Active now / ongoing"),
    ("SAP Business One Consultant", "Amaris Consulting (client undisclosed)", "BE", "Herstal", "SAP Business One", SOL_B1, PT_RUN, "Active now / ongoing"),
    ("SAP Application Security Consultant", "BDO Belgium", "BE", "Zaventem", "SAP ECC/S/4HANA Security & GRC", SOL_SEC, PT_SEC, "Active now / ongoing"),
    ("SAP Global Trade Services Consultant", "delaware BeLux", "BE", "Ghent", "SAP GTS", SOL_GTS, PT_IMPL, "Active now / ongoing"),
    ("SAP Enterprise Architect", "KPMG Belgium", "BE", "Zaventem", "S/4HANA + BTP Integration", SOL_S4, PT_ARCH, "Active now / ongoing"),
    ("SAP Archiving & ILM Manager", "Deloitte Belgium", "BE", "Zaventem", "SAP Data Archiving / ILM", SOL_DATA, PT_SEC, "Active now / ongoing"),
    ("SAP EWM/WM Consultant", "Accenture Belgium", "BE", "Brussels", "SAP EWM / WM", SOL_WM, PT_IMPL, "Active now / ongoing"),
    ("SAP S4HANA Functional Analyst & Developer", "Talan (client undisclosed)", "BE", "Brussels", "S/4HANA (ECC migration)", SOL_S4, PT_MIG, "Active now / ongoing"),
    ("SAP Developer", "Kingfisher Recruitment (client undisclosed)", "BE", "Oud-Turnhout", "SAP ECC -> future S/4HANA", SOL_DEV, PT_MIG, "Upcoming migration (no date stated)"),
    ("SAP Solution Architect", "Sopra Steria", "BE", "Brussels", "SAP (general architecture, multi-client)", SOL_S4, PT_ADV, "Active now / ongoing"),
    ("SAP PS Senior Manager", "Accenture Belgium", "BE", "Brussels", "SAP PS + RISE with SAP / S/4HANA", SOL_S4, PT_ARCH, "Active now / ongoing"),
    ("Functional Analyst SAP Logistics (Sr)", "John Cockerill", "BE", "Seraing", "S/4HANA Logistics (PS/MM/SD/PP/PM/QM)", SOL_PP, PT_MIG, "Active transformation, ongoing"),
    ("SAP Supply Chain Planning Specialist", "Deloitte Belgium", "BE", "Zaventem", "SAP IBP / PP-DS", SOL_PP, PT_IMPL, "Active now / ongoing"),
    ("SAP Specialist", "Indaver", "BE", "Beveren", "S/4HANA (post-migration)", SOL_S4, PT_RUN, "Post-migration, ongoing"),
    ("SAP Sales", "Match Profiler (client undisclosed)", "BE", "Kaprijke", "SAP SD & MM", SOL_SD, PT_IMPL, "Active now / ongoing"),
    ("Lead Consultant – SAP Manufacturing", "Emixa (client undisclosed)", "BE", "Antwerp", "SAP PP / PP-PI / PP-DS", SOL_PP, PT_IMPL, "Active now / ongoing"),
    ("SAP Solutions Specialist", "Deloitte Belgium", "BE", "Zaventem", "S/4HANA (multi-module)", SOL_S4, PT_IMPL, "Active now / ongoing"),
    ("SAP Team Lead", "Crop's", "BE", "West Flanders", "SAP SD/MM + Clean Core/BTP", SOL_SD, PT_MIG, "Active transition, ongoing"),
    ("SAP-PP & MES Consultant", "stow Group", "BE", "Spiere-Helkijn", "SAP PP + MES", SOL_PP, PT_IMPL, "Active MES rollout, ongoing"),
    ("Consultant(e) SAP S/4HANA (M/F)", "mc2i (client undisclosed)", "BE", "Brussels", "S/4HANA (Greenfield/Brownfield/Selective)", SOL_S4, PT_IMPL, "Active now / ongoing"),
    ("Solution Architect SAP", "Madison Recruitment (client undisclosed)", "BE", "Genk", "SAP EWM + S/4HANA", SOL_WM, PT_MIG, "Active now / ongoing"),
    ("SAP PP Consultant", "delaware BeLux", "BE", "Flemish Region", "S/4HANA PP", SOL_PP, PT_IMPL, "Active now / ongoing"),
    ("SAP S/4HANA Finance Consultant", "Eursap (client undisclosed)", "Remote (France)", "Remote", "S/4HANA Finance", SOL_FI, PT_IMPL, "Explicit start date: 5 Oct 2026"),
    ("SAP BTP Techno-Functional Consultant", "Eursap (client undisclosed)", "Remote (France)", "Remote", "SAP BTP", SOL_BTP, PT_IMPL, "Explicit start date: 5 Oct 2026"),
    ("SAP Project Manager / Trainer", "Eursap (client undisclosed)", "Remote", "Remote", "SAP (general PM/Training)", SOL_AI, PT_PM, "Explicit start date: 1 Sep 2026"),
]

HEADER = ["SAP Solution (category)", "Solution Detail", "Project Type", "Time Frame",
          "Company", "Country", "Location", "Role Hiring"]
HEADER_FILL = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def build_forecast_sheet(ws):
    ws.title = "Forecast by company"
    ws.append(HEADER)
    style_header(ws)
    for role, company, country, location, detail, category, ptype, timeframe in sorted(
        FORECAST, key=lambda r: (r[5], r[1])
    ):
        ws.append([category, detail, ptype, timeframe, company, country, location, role])
    autosize(ws, [30, 34, 20, 34, 32, 14, 20, 46])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"


def build_summary_sheet(ws, key_index, title, col_label, widths):
    ws.title = title
    ws.append([col_label, "Open Roles", "Companies", "Companies (\"client undisclosed\" = via agency)"])
    style_header(ws)

    grouped = defaultdict(list)
    for row in FORECAST:
        grouped[row[key_index]].append(row[4])  # company name

    rows = []
    for key, companies in grouped.items():
        uniq = sorted(set(companies))
        rows.append((key, len(companies), len(uniq), "; ".join(uniq)))
    rows.sort(key=lambda r: r[1], reverse=True)
    for row in rows:
        ws.append(list(row))
    autosize(ws, widths)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"


def build_readme_sheet(ws):
    ws.title = "Read me"
    ws.append(["SAP Demand Forecast — Belgium & Netherlands: methodology and limits"])
    ws["A1"].font = Font(bold=True, size=13)
    lines = [
        "",
        f"Generated {REPORT_DATE}, derived from the same 65 live SAP job postings gathered",
        "in sap_jobs_be_nl_2026-08-06.xlsx (30 Netherlands, 35 Belgium roles, LinkedIn-",
        "sourced, posted within 7 days of the pull date).",
        "",
        "WHAT THIS IS: a structured read-out of current hiring signals, grouped by SAP",
        "solution category, project type, and stated/inferred time frame. Useful for",
        "spotting which solutions and project types are generating the most open roles",
        "right now, and which companies are behind them.",
        "",
        "WHAT THIS IS NOT: a statistical or predictive forecast. There is no trend model,",
        "no historical time series, and no market-wide survey behind it — it is a",
        "classification of one snapshot of job ads. Two limits to keep in mind:",
        "  1. Sample size: 65 roles from one search pass, not the full BE/NL SAP market.",
        "  2. Time frame accuracy: only 3 of 65 roles stated an explicit start date or",
        "     engagement length (the 2 Eursap contract roles, and Robert Half's stated",
        "     ~4-year Ghent engagement). All other rows read \"Active now / ongoing\" —",
        "     that only means the company is hiring today, not when their underlying SAP",
        "     project starts, ends, or how long it will run.",
        "",
        "CATEGORIZATION IS A JUDGMENT CALL: \"Solution Detail\" is close to the posting's",
        "own words; \"Solution (category)\" and \"Project Type\" are the author's bucketing",
        "into a smaller set so the summary sheets can aggregate meaningfully — e.g. a role",
        "mentioning both SD and MM was assigned to whichever the posting emphasized more.",
        "Cross-check the Detail column or the original posting (see",
        "sap_jobs_be_nl_2026-08-06.xlsx) if precision matters for a decision.",
        "",
        "COMPANY NAMES: where a job was posted by a recruitment agency/consultancy on",
        "behalf of an undisclosed end client (Robert Half, Madison Recruitment, RED Global,",
        "Match Profiler, Kingfisher Recruitment, ThoughtLabs Belgium, JoBBsquare, Aures,",
        "Talan, mc2i, Panda International, Amaris Consulting), the row is labeled",
        "\"<agency> (client undisclosed)\" — the true end-client company is not public.",
    ]
    for line in lines:
        ws.append([line])
    ws.column_dimensions["A"].width = 92
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_xlsx():
    wb = Workbook()
    build_forecast_sheet(wb.active)
    build_summary_sheet(wb.create_sheet(), key_index=5, title="Solution demand summary",
                         col_label="SAP Solution (category)", widths=[36, 12, 12, 70])
    build_summary_sheet(wb.create_sheet(), key_index=6, title="Project type summary",
                         col_label="Project Type", widths=[26, 12, 12, 70])
    build_readme_sheet(wb.create_sheet())
    wb.save(str(XLSX_OUTPUT_PATH))
    print(f"Forecast XLSX saved to: {XLSX_OUTPUT_PATH}")


if __name__ == "__main__":
    build_xlsx()

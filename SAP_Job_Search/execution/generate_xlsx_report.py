#!/usr/bin/env python3
"""
Builds the SAP Job Search workbook: one row per open role with the posting
party's contact details (where publicly available), following the columns
requested:
  1. Function            2. Procurement party / internal job
  3. Contact person       4. Contact details
  5. When posted          6. Location

Data is a snapshot transcribed from `SAP_Job_Search/findings/sap_jobs_be_nl_<date>.md`.
Where no named recruiter or direct phone/email is publicly listed for a given
company, that is stated plainly rather than invented — most large direct
employers (Accenture, Deloitte, Capgemini, ...) don't publish a named
recruiter per vacancy; you apply via their careers portal and a recruiter is
assigned afterwards. True third-party staffing/recruitment agencies (Eursap,
Madison Recruitment, Robert Half, ...) do publish a general desk you can call.
"""

import sys
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPORT_DATE = "2026-08-06"
root_dir = Path(__file__).parent.parent
findings_dir = root_dir / "findings"
findings_dir.mkdir(exist_ok=True, parents=True)
XLSX_OUTPUT_PATH = findings_dir / f"sap_jobs_be_nl_{REPORT_DATE}.xlsx"

NOT_PUBLISHED = "Not published — apply via company careers page"

# company -> (party_type, contact_person, contact_details)
# party_type is "Agency" (third-party recruiter/staffing procurement party) or
# "Direct" (the hiring company itself; no intermediary procurement party).
CONTACTS = {
    "Eursap": ("Agency", "Recruitment desk (no named individual published)",
               "BE +32 (0)2 808 5964 / NL +31 (0)20 890 8064 / info@eursap.eu"),
    "Michael Page": ("Agency", "Recruitment desk (no named individual published)",
                      "Brussels +32 (0)2 509 4545 / Amsterdam +31 (0)20 578 9444"),
    "Hays": ("Agency", "Recruitment desk (no named individual published)",
             "Amsterdam +31 (0)20 363 0310"),
    "Robert Half": ("Agency", "Recruitment desk (no named individual published)",
                     "Belgium +32 3 241 14 28"),
    "Approach People Recruitment": ("Agency", "Recruitment desk (no named individual published)",
                                     "Group line +353 1 400 3500"),
    "GS-IT Recruitment": ("Agency", "Recruitment desk (no named individual published)",
                           "Contact form only — gs-it-recruitment.eu"),
    "YER": ("Agency", "Recruitment desk (no named individual published)",
            "Branch locator — yer.nl"),
    "Madison Recruitment": ("Agency", "Tuur Vandeurzen",
                             "+32 473 40 69 30 / t.vandeurzen@madison.be"),
    "Kingfisher Recruitment": ("Agency", "Recruitment desk (no named individual published)",
                                "+32 14 89 98 51 (Herentals office)"),
    "RED Global": ("Agency", "Recruitment desk (no named individual published)",
                    "+31 6 47 30 34 19 (Rotterdam office)"),
    "Match Profiler": ("Agency", "Recruitment desk (no named individual published)",
                        "+351 21 330 4403 / emprego@m-profiler.com — Portugal HQ, "
                        "BE branch contact unconfirmed this pass"),
    "Accenture NL": ("Direct", "—", NOT_PUBLISHED + " (accenture.com/nl-en/careers)"),
    "Accenture Belgium": ("Direct", "—",
                           "+32 (0)2 226 72 11 (Brussels office, general switchboard) — "
                           "careers via accenture.com/be-en/careers"),
    "Deloitte": ("Direct", "—", NOT_PUBLISHED + " (careersatdeloitte.com/contact)"),
    "Deloitte Belgium": ("Direct", "—",
                          "+32 2 600 60 00 (Zaventem HQ, general switchboard) — "
                          "careers via deloitte.com/be/en/careers"),
    "Capgemini": ("Direct", "—", NOT_PUBLISHED + " (capgemini.com/nl-nl/contact-us)"),
    "delaware the Netherlands": ("Direct", "—", "+32 56 27 44 44 / info@delaware.pro"),
    "delaware BeLux": ("Direct", "—", "+32 56 27 44 44 / info@delaware.pro"),
    "NTT DATA Business Solutions": ("Direct", "—",
                                     NOT_PUBLISHED + " (nttdata-solutions.com/bnl/contact) — "
                                     "a recruiter first-named \"Sharon\" was referenced online "
                                     "without a verifiable surname/direct line"),
    "Sopra Steria": ("Direct", "—", "+32 2 208 72 72 / +32 2 566 66 66 (Brussels)"),
    "KPMG Belgium": ("Direct", "—", "+32 2 708 4300 (Zaventem office, general switchboard)"),
    "BCG Platinion": ("Direct", "—", "+31 20 548 4000 (Amsterdam office, general switchboard)"),
    "Stedin": ("Direct", "—", NOT_PUBLISHED + " (werkenbij.stedin.net/contact — "
               "recruiter shown only once you open the specific vacancy)"),
}


def contact(company):
    party_type, person, details = CONTACTS.get(
        company, ("Direct", "—", NOT_PUBLISHED + " — not individually researched this pass")
    )
    label = f"{company} ({'agency' if party_type == 'Agency' else 'direct employer'})"
    return label, person, details


# (title, company, location, posted) — same source data as generate_pdf_report.py
NL_ROLES = [
    ("SAP Integration Specialist", "Alliander", "Arnhem", "8 min ago (6 Aug 2026)"),
    ("SAP S/4HANA Roadmap & Business Transformation Advisor", "Accenture NL", "Amsterdam", "15 hrs ago (5 Aug 2026)"),
    ("SAP Solution Architect", "Brabant Water N.V.", "'s-Hertogenbosch", "16 hrs ago (5 Aug 2026)"),
    ("SAP-ontwikkelaar", "myBrand | Conclusion", "Maarssen", "18 hrs ago (5 Aug 2026)"),
    ("SAP/BTP Development Specialist", "Nyrstar", "Budel", "20 hrs ago (5 Aug 2026)"),
    ("SAP CAP Developer", "delaware the Netherlands", "Den Bosch", "1 day ago (5 Aug 2026)"),
    ("SAP S/4HANA Data Transformation Advisor", "Accenture NL", "Amsterdam", "2 days ago (4 Aug 2026)"),
    ("Manager SAP S/4HANA Sales and Distribution", "Deloitte", "Amsterdam", "2 days ago (4 Aug 2026)"),
    ("SAP Production Consultant", "NTT DATA Business Solutions", "'s-Hertogenbosch", "2 days ago (4 Aug 2026)"),
    ("SAP Integration Consultant", "Accenture NL", "Amsterdam", "4 days ago (2 Aug 2026)"),
    ("SAP Sales Distribution Consultant (SD/OTC)", "Accenture NL", "Amsterdam", "4 days ago (2 Aug 2026)"),
    ("SAP Development Architect & Team Lead", "Accenture NL", "Amsterdam", "5 days ago (1 Aug 2026)"),
    ("SAP Architect – NL", "Capgemini", "Utrecht", "5 days ago (1 Aug 2026)"),
    ("SAP Manufacturing Manager", "Accenture NL", "Amsterdam", "5 days ago (1 Aug 2026)"),
    ("SAP S/4 Digital EAM Consultant", "Accenture NL", "Amsterdam", "5 days ago (1 Aug 2026)"),
    ("SAP Security Lead", "Scandinavian Tobacco Group", "Waalre", "5 days ago (1 Aug 2026)"),
    ("Super User SAP", "Scania Production & Logistics NL", "Hasselt (Overijssel)", "5 days ago (1 Aug 2026)"),
    ("SAP Group Reporting & Public Cloud", "RED Global", "South Holland", "5 days ago (1 Aug 2026)"),
    ("SAP developer", "Stedin", "Delft", "6 days ago (31 Jul 2026)"),
    ("Technical Application manager", "Panda International", "Bilthoven", "6 days ago (31 Jul 2026)"),
    ("Consultant SAP EAM", "Deloitte", "Amsterdam", "6 days ago (31 Jul 2026)"),
    ("SAP Sourcing & Procurement Consultant", "Accenture NL", "Amsterdam", "6 days ago (31 Jul 2026)"),
    ("SAP ABAP Developer", "Stedin", "Delft", "1 week ago (30 Jul 2026)"),
    ("SAP C4C consultant", "Stedin", "Rotterdam", "1 week ago (30 Jul 2026)"),
    ("AI Tech Consultant (ERP/SAP)", "BCG Platinion", "Amsterdam", "1 week ago (30 Jul 2026)"),
    ("Lead AI Tech Architect (ERP/SAP)", "BCG Platinion", "Amsterdam", "1 week ago (30 Jul 2026)"),
    ("Senior AI Tech Architect (ERP/SAP)", "BCG Platinion", "Amsterdam", "1 week ago (30 Jul 2026)"),
    ("Senior SAP Developer", "vidaXL", "Venlo", "1 week ago (30 Jul 2026)"),
    ("SAP Developer", "Applied Medical", "Amersfoort", "1 week ago (30 Jul 2026)"),
    ("SAP Treasury Risk Management", "Accenture NL", "Amsterdam", "1 week ago (30 Jul 2026)"),
]

BE_ROLES = [
    ("SAP Analyst", "ThoughtLabs Belgium", "Brussels", "55 min ago (6 Aug 2026)"),
    ("SAP PP/QM", "JoBBsquare België", "Antwerp", "8 hrs ago (6 Aug 2026)"),
    ("SAP Functioneel Analyst", "Madison Recruitment", "Aalst", "13 hrs ago (5 Aug 2026)"),
    ("SAP GTS Consultant", "Deloitte Belgium", "Zaventem", "16 hrs ago (5 Aug 2026)"),
    ("SAP TM Consultant", "Deloitte Belgium", "Zaventem", "18 hrs ago (5 Aug 2026)"),
    ("System Exploitation Engineer – SAP BC Member", "Sibelga", "Brussels", "1 day ago (5 Aug 2026)"),
    ("Functioneel SAP-analist", "Aures", "Wingene", "1 day ago (5 Aug 2026)"),
    ("SAP Enterprise Architect", "Robert Half", "Ghent", "1 day ago (5 Aug 2026)"),
    ("SAP BDC Lead", "KPMG Belgium", "Zaventem", "1 day ago (5 Aug 2026)"),
    ("Senior SAP Sales & Distribution Consultant", "delaware BeLux", "Ghent", "1 day ago (5 Aug 2026)"),
    ("Responsable Fonctionnel SAP Quality / WMS", "Safran", "Liège", "2 days ago (4 Aug 2026)"),
    ("Responsable Fonctionnel SAP Finance", "Safran", "Liège", "2 days ago (4 Aug 2026)"),
    ("SAP Architect (SD/MM)", "Amon", "Kaprijke", "2 days ago (4 Aug 2026)"),
    ("SAP FICO Consultant", "Deloitte Belgium", "Zaventem", "2 days ago (4 Aug 2026)"),
    ("SAP Business One Consultant", "Amaris Consulting", "Herstal", "2 days ago (4 Aug 2026)"),
    ("SAP Application Security Consultant", "BDO Belgium", "Zaventem", "2 days ago (4 Aug 2026)"),
    ("SAP Global Trade Services Consultant", "delaware BeLux", "Ghent", "2 days ago (4 Aug 2026)"),
    ("SAP Enterprise Architect", "KPMG Belgium", "Zaventem", "3 days ago (3 Aug 2026)"),
    ("SAP Archiving & ILM Manager", "Deloitte Belgium", "Zaventem", "3 days ago (3 Aug 2026)"),
    ("SAP EWM/WM Consultant", "Accenture Belgium", "Brussels", "4 days ago (2 Aug 2026)"),
    ("SAP S/4HANA Functional Analyst & Developer", "Talan", "Brussels", "4 days ago (2 Aug 2026)"),
    ("SAP Developer", "Kingfisher Recruitment", "Oud-Turnhout", "5 days ago (1 Aug 2026)"),
    ("SAP Solution Architect", "Sopra Steria", "Brussels", "5 days ago (1 Aug 2026)"),
    ("SAP PS Senior Manager", "Accenture Belgium", "Brussels", "5 days ago (1 Aug 2026)"),
    ("Functional Analyst SAP Logistics (Sr)", "John Cockerill", "Seraing", "5 days ago (1 Aug 2026)"),
    ("SAP Supply Chain Planning Specialist", "Deloitte Belgium", "Zaventem", "6 days ago (31 Jul 2026)"),
    ("SAP Specialist", "Indaver", "Beveren", "6 days ago (31 Jul 2026)"),
    ("SAP Sales", "Match Profiler", "Kaprijke", "1 week ago (30 Jul 2026)"),
    ("Lead Consultant – SAP Manufacturing", "Emixa", "Antwerp", "1 week ago (30 Jul 2026)"),
    ("SAP Solutions Specialist", "Deloitte Belgium", "Zaventem", "1 week ago (30 Jul 2026)"),
    ("SAP Team Lead", "Crop's", "West Flanders", "1 week ago (30 Jul 2026)"),
    ("SAP-PP & MES Consultant", "stow Group", "Spiere-Helkijn", "1 week ago (30 Jul 2026)"),
    ("Consultant(e) SAP S/4HANA (M/F)", "mc2i", "Brussels", "1 week ago (30 Jul 2026)"),
    ("Solution Architect SAP", "Madison Recruitment", "Genk", "1 week ago (30 Jul 2026)"),
    ("SAP PP Consultant", "delaware BeLux", "Flemish Region", "1 week ago (30 Jul 2026)"),
]

# (title, company) -> (job description summary, [requirement bullets])
# Sourced from each role's individual LinkedIn posting (fetched same pass as the role
# list itself). Condensed to the essentials — see the original posting via LinkedIn for
# the full text.
JOB_DETAILS = {
    ("SAP Integration Specialist", "Alliander"): (
        "Build and maintain integrations between SAP and Alliander's digital landscape, "
        "supporting secure/efficient data flow for the Dutch energy grid.",
        ["Applied sciences degree or equivalent experience", "4+ yrs system integrations/application interfaces",
         "Hands-on SAP BTP Integration Suite, API Management, Cloud Integration", "Fluent Dutch and English",
         "Plus: SAP Build Process Automation, Terraform, Event Mesh, Kafka, Groovy"]),
    ("SAP S/4HANA Roadmap & Business Transformation Advisor", "Accenture NL"): (
        "Guide clients through S/4HANA transformations: develop strategic roadmaps, lead "
        "blueprinting, translate business requirements into system designs (all levels).",
        ["Master's in IT/Business or bachelor's + experience", "2-10+ yrs SAP consulting, S/4HANA roadmaps/strategy",
         "Track record leading transformations vision-to-go-live", "SAP Activate methodology, blueprint docs",
         "Functional breadth across Finance/Logistics/Sales/Procurement", "Willingness to travel"]),
    ("SAP Solution Architect", "Brabant Water N.V."): (
        "Design and oversee scalable SAP solutions for asset & supply chain management, "
        "bridging business strategy and technical execution for a drinking-water utility.",
        ["HBO/WO education", "7+ yrs as SAP specialist (asset mgmt/supply chain), 2+ yrs architecture/lead role",
         "Solution architecture in complex enterprise environments", "Fluent Dutch (mandatory)",
         "Plus: Azure DevOps, Scrum, SAP EAM/SCM certifications"]),
    ("SAP-ontwikkelaar", "myBrand | Conclusion"): (
        "Help clients modernize SAP solutions via cloud-focused development; maintain "
        "existing systems and guide teams toward contemporary SAP technologies.",
        ["Extensive ABAP Cloud Development model experience", "Ability to design cloud-oriented SAP solutions",
         "Knowledge of SAP Cloud ERP / Cloud ERP Private", "Senior-level, strong stakeholder communication",
         "Mentoring capability"]),
    ("SAP/BTP Development Specialist", "Nyrstar"): (
        "Design and maintain applications for Metal Accounting, Exposure Management and "
        "Finance; modernize custom code and act as internal SME.",
        ["Bachelor's/Master's in CS, Information Systems or Engineering", "SAP ABAP certification required; BTP/BW preferred",
         "3-5+ yrs SAP ABAP development", "SAP integration tech: OData, APIs, IDocs, RFCs",
         "SAP BTP (integration, automation, GenAI) a plus", "Fluent English; NL/FR/DE valued"]),
    ("SAP CAP Developer", "delaware the Netherlands"): (
        "Build SAP Cloud Application Programming solutions with TypeScript/React/Fiori in "
        "an event-driven architecture; mentor juniors, own client engagements.",
        ["2-4 yrs TypeScript/JavaScript, SAP CAP, Fiori, React", "Event-driven architecture/integration experience preferred",
         "SAP HANA and PostgreSQL knowledge", "Git, CI/CD, Agile familiarity", "Dutch and English fluency"]),
    ("SAP S/4HANA Data Transformation Advisor", "Accenture NL"): (
        "Guide clients through S/4HANA data strategy, governance frameworks and analytics "
        "enablement across Finance, Supply Chain and HR (all levels).",
        ["2-10+ yrs SAP data domains (MDM, DW, Analytics, or S/4 migration)", "Master's/bachelor's + experience",
         "SAP Activate + blueprint workshop facilitation", "Plus: SAP MDG, BW/4HANA, Analytics Cloud, Datasphere",
         "Willingness to travel"]),
    ("Manager SAP S/4HANA Sales and Distribution", "Deloitte"): (
        "Lead SAP S/4HANA Sales & Distribution implementations for global clients, from "
        "assessment through go-live, guiding design and mentoring teams.",
        ["Bachelor's or Master's degree", "6+ yrs S/4HANA SD implementation experience", "4+ completed implementations, exploration-to-go-live",
         "Knowledge of CX/e-commerce/CPQ and/or BRIM service processes", "Fluent English and Dutch"]),
    ("SAP Production Consultant", "NTT DATA Business Solutions"): (
        "Optimize manufacturing/supply chain processes using SAP PP/DS, APO, IBP for Food & "
        "Agri, Life Sciences and Discrete Manufacturing clients.",
        ["HBO or WO degree", "2+ yrs SAP PP, PP/DS, APO and/or IBP in complex projects", "Full-cycle implementation/customization expertise",
         "Dutch and English proficiency", "32+ hrs/week, hybrid in 's-Hertogenbosch"]),
    ("SAP Integration Consultant", "Accenture NL"): (
        "Design/implement enterprise integration solutions (S/4HANA, SAP Cloud Platform) "
        "for multinational clients; gather requirements, build interfaces.",
        ["Bachelor's/Master's + 5-10 yrs as SAP integration consultant", "CPI, PI/PO, API Management experience",
         "Cloud integration knowledge preferred", "Fluent Dutch and English", "Self-driven, team player"]),
    ("SAP Sales Distribution Consultant (SD/OTC)", "Accenture NL"): (
        "Implement SAP sales/customer experience solutions across the full lifecycle — "
        "order-to-cash, billing, pricing, logistics execution.",
        ["Bachelor's or master's degree", "2+ yrs SAP OTC, CX or SD modules", "2+ full-cycle projects completed",
         "Excellent written/oral English", "Plus: S/4HANA certification, Dutch language"]),
    ("SAP Development Architect & Team Lead", "Accenture NL"): (
        "Design and govern modern SAP solutions on S/4HANA and SAP BTP, establishing "
        "architectural standards aligned with SAP's Clean Core principles.",
        ["Bachelor's/Master's in IT + hands-on S/4HANA development", "ABAP (OOP), ABAP Cloud, RAP, CDS Views, OData, UI5, Fiori, CAP, BTP",
         "Clean Core compliance, Business Application Studio", "Proven team leadership", "Fluent English and Dutch; EU citizenship; travel"]),
    ("SAP Architect – NL", "Capgemini"): (
        "Lead complex digital transformations for major Dutch organizations, designing "
        "end-to-end SAP solutions and guiding transitions to S/4HANA Cloud.",
        ["8+ yrs SAP, Solution/Enterprise Architect or Lead Consultant background", "Deep S/4HANA, Cloud ERP, integrations/extensions knowledge",
         "SAP Activate, Clean Core, Best Practices", "HBO or university degree", "Fluent Dutch and English"]),
    ("SAP Manufacturing Manager", "Accenture NL"): (
        "Lead client transformations using SAP S/4HANA and digital manufacturing "
        "capabilities across automotive, life sciences and consumer goods.",
        ["Bachelor's/Master's in Engineering, Manufacturing, IT or related", "8+ yrs SAP Manufacturing/PP, ideally S/4HANA",
         "Shop-floor process understanding (execution, MRP)", "Experience leading global implementation teams",
         "Fluent English; Dutch or other EU languages preferred"]),
    ("SAP S/4 Digital EAM Consultant", "Accenture NL"): (
        "Entry-level EAM consulting: design/implement Enterprise Asset Management "
        "solutions on S/4HANA — blueprinting, configuration, testing, go-live.",
        ["Bachelor's/Master's in Engineering, IT, Business or related", "Fluent English and Dutch (or actively learning)",
         "SAP PM(CS)/S/4HANA EAM experience preferred", "Strong problem-solving, willingness to learn"]),
    ("SAP Security Lead", "Scandinavian Tobacco Group"): (
        "Oversee security strategy across the S/4HANA landscape: governance frameworks, "
        "authorizations/access controls, regulatory compliance.",
        ["Bachelor's in IT, CS, Cybersecurity or related", "3+ yrs SAP Security/Authorization administration",
         "S/4HANA security, Fiori, role design methodologies", "SoD controls and audit frameworks knowledge",
         "Hybrid, permanent, 40 hrs/week in Waalre"]),
    ("Super User SAP", "Scania Production & Logistics NL"): (
        "Optimize logistical processes within the Logistic Center Engineering team, "
        "combining IT expertise with warehouse operations (EWM in S/4HANA).",
        ["Bachelor's in logistics, IT or technical business studies", "Demonstrable SAP EWM in S/4HANA experience",
         "Warehouse/logistics environment experience", "Fluent Dutch and English", "€4,000–€5,500 gross/month"]),
    ("SAP Group Reporting & Public Cloud", "RED Global"): (
        "Contract role via recruiter RED Global: Senior SAP Group Reporting & "
        "Consolidation Expert for an international client.",
        ["Extensive hands-on SAP Group Reporting / financial consolidation experience", "Strong grasp of consolidation processes/reporting methodologies",
         "Proficient spoken/written English", "Senior level"]),
    ("SAP developer", "Stedin"): (
        "Build and enhance SAP systems (ABAP, Fiori/UI5, integrations) powering the "
        "energy grid, in a DevOps team, hybrid in Delft.",
        ["HBO/WO in (Business) IT or Software Engineering", "2+ yrs proven SAP Developer experience",
         "SAP ABAP, Integration Suite, Fiori/UI5 or SAP Build basics", "Native-level Dutch (C2)",
         "€4,045–€5,623/month, 32-40 hr week"]),
    ("Technical Application manager", "Panda International"): (
        "Manage/optimize a life-sciences SAP ERP landscape; bridge business "
        "stakeholders and IT in a GMP-regulated environment.",
        ["5+ yrs SAP ERP in complex organizations", "SAP ECC and/or S/4HANA knowledge",
         "Experience with SAP modules PP, MM, SD, FI/CO etc.", "Incident/change/release management expertise",
         "Plus: pharma/biotech, SAP validation/compliance"]),
    ("Consultant SAP EAM", "Deloitte"): (
        "Help global organizations transform asset management via modern SAP EAM "
        "solutions, bridging business and IT across international projects.",
        ["Bachelor's/Master's in Supply Chain, Industrial Engineering or comparable", "3+ yrs hands-on SAP EAM consultant, configuration expertise",
         "1+ full implementation cycle, Agile familiarity", "Fluent Dutch and English",
         "Plus: SAP Cloud ERP, Intelligent Asset Management"]),
    ("SAP Sourcing & Procurement Consultant", "Accenture NL"): (
        "Implement procurement solutions (S/4HANA, Ariba) within Source to Pay for "
        "international clients across blueprint, design, realization, testing.",
        ["Bachelor's/Master's degree", "5+ yrs SAP implementation project experience", "Strong Source to Pay process knowledge",
         "Fluent Dutch and English", "Plus: S/4HANA/Ariba/Coupa certification"]),
    ("SAP ABAP Developer", "Stedin"): (
        "Build/optimize SAP solutions for the supply chain and energy transition in an "
        "Agile DevOps team, using ABAP, RAP, Fiori, UI5, CPI.",
        ["HBO-level, 5 yrs as SAP Developer (3+ yrs ABAP/RAP/Fiori/UI5/CPI)", "SAP PI/PO integrations, APIs, SAP BTP experience",
         "Agile/Scrum, SAFe knowledge", "Fluent Dutch (C1/C2)", "€4,548–€7,267/month, hybrid in Delft"]),
    ("SAP C4C consultant", "Stedin"): (
        "Develop/optimize the CRM system (SAP C4C) within a DevOps team; configure "
        "modules and translate business processes into technical solutions.",
        ["HBO-level education, fluent Dutch (C2)", "3+ yrs functional consultant/developer in SAP C4C with CRM background",
         "Agile methodology proficiency", "Up to €7,124.60/month gross (40 hrs)"]),
    ("AI Tech Consultant (ERP/SAP)", "BCG Platinion"): (
        "Support ERP-enabled business transformation programs, combining hands-on "
        "delivery with strategic consulting skill development.",
        ["3-6 yrs ERP systems experience", "Consulting/Systems Integrator background, 1+ SAP module",
         "Familiarity with AI-enabled productivity tools (GenAI)", "Client-facing PMO/delivery/change management exposure",
         "STEM or business degree (MBA/MSc advantageous)"]),
    ("Lead AI Tech Architect (ERP/SAP)", "BCG Platinion"): (
        "Design/oversee large-scale SAP transformations, defining technical strategy and "
        "integrating AI capabilities into business outcomes.",
        ["10+ yrs SAP architecture/delivery, consulting background", "S/4HANA adoption, multi-module design, integration architecture (BTP, Integration Suite)",
         "Master data governance, BW/4HANA, DataSphere, SAC", "Ability to advise C-suite", "Willingness to travel internationally"]),
    ("Senior AI Tech Architect (ERP/SAP)", "BCG Platinion"): (
        "Lead large-scale ERP transformations, designing SAP target landscapes and "
        "integration strategies, identifying AI-driven value.",
        ["6+ yrs SAP architecture/delivery experience", "Multi-module solution design, S/4HANA adoption strategy",
         "SAP BTP, Integration Suite, middleware skills", "Master data governance/analytics platform design",
         "STEM or business degree; willing to travel worldwide"]),
    ("Senior SAP Developer", "vidaXL"): (
        "Lead solution design/development/deployment in the S/4HANA environment for "
        "logistics and supply chain applications.",
        ["7+ yrs SAP Developer/Technical Lead in S/4HANA", "ABAP, RAP, SAP development best practices",
         "SAP BTP, Fiori, OData, CDS views", "DevOps, CI/CD, Git-based development"]),
    ("SAP Developer", "Applied Medical"): (
        "Design/develop/maintain SAP solutions (ABAP, Fiori) for the European HQ in "
        "Amersfoort, collaborating with technical and functional stakeholders.",
        ["3+ yrs ABAP programming (user-exits, ALV, Web Dynpro, BAPI, BAdIs, Smartforms)", "Modern ABAP (7.40+), S/4HANA, SQL",
         "Bachelor's in computer science or equivalent", "SAP Fiori development", "Plus: RAP, Fiori Elements, BTP, cGMP, SuccessFactors"]),
    ("SAP Treasury Risk Management", "Accenture NL"): (
        "Manager/Senior Manager leading treasury/finance transformation projects across "
        "EMEA, combining SAP Treasury functional expertise with S/4HANA implementation.",
        ["Strong SAP Finance/Treasury consulting background", "SAP Treasury (BCM, MBC, Cash Mgmt, Risk, Hedging)",
         "Debt, investments, FX hedging, in-house banking knowledge", "S/4HANA Treasury suite familiarity",
         "University finance credentials or treasury certifications"]),
    ("SAP Analyst", "ThoughtLabs Belgium"): (
        "SAP Materials Management analyst role in Brussels with heavy public-sector "
        "focus, Associate level.",
        ["Dutch or French at C1; English at B2", "10+ yrs advanced SAP MM expertise", "5+ yrs each SAP Ariba Commerce Automation and Sourcing",
         "5+ yrs SAP EDI knowledge", "15+ yrs public sector background"]),
    ("SAP PP/QM", "JoBBsquare België"): (
        "Manage production execution, quality management and production planning for a "
        "manufacturing environment, integrating shop floor and LIMS systems.",
        ["SAP Production Execution (PE) expertise, EWM integration a plus", "SAP Quality Management (QM), Labvantage/LIMS integration knowledge",
         "PPDS understanding in manufacturing context", "Mid-senior level, English"]),
    ("SAP Functioneel Analyst", "Madison Recruitment"): (
        "Analyze/optimize supply chain processes and configure SAP for an international "
        "production company in Aalst, via Madison Recruitment.",
        ["5+ yrs SAP experience, preferably SD module", "Strong communication/stakeholder management",
         "Analytical mindset, problem-solving", "System configuration, testing, documentation experience",
         "3 WFH days/week offered"]),
    ("SAP GTS Consultant", "Deloitte Belgium"): (
        "Entry-level consulting: design/implement trade compliance and customs "
        "solutions (SAP GTS) for clients across industries.",
        ["2+ yrs professional SAP GTS experience", "Global trade regulations/customs/compliance knowledge",
         "1+ implementation supported", "Bachelor's degree, fluent English", "~20% travel; ABAP/SAP PI a plus"]),
    ("SAP TM Consultant", "Deloitte Belgium"): (
        "Support Transportation Management implementations across industries in "
        "Deloitte's Center of Excellence, entry-level consultant.",
        ["2-3 yrs hands-on SAP TM implementation, 1+ end-to-end project", "Order Mgmt, Freight Planning/Execution/Settlement expertise",
         "University degree in Supply Chain, Logistics, CS or Business", "English professional; French or Dutch preferred"]),
    ("System Exploitation Engineer – SAP BC Member", "Sibelga"): (
        "Manage technical operations/maintenance of enterprise SAP systems: production "
        "deployments, technical analysis, platform evolution.",
        ["15+ yrs operational SAP experience in large enterprises", "Bachelor's/Master's in IT",
         "Fluent French or Dutch + technical English", "Deep SAP NetWeaver (ABAP/Java), Basis, S/4HANA, Oracle/HANA expertise",
         "ITIL Foundation recommended; evening/weekend deployment availability"]),
    ("Functioneel SAP-analist", "Aures"): (
        "Bridge business and IT for a manufacturing organization in Wingene: translate "
        "operational needs into digital (SAP) solutions.",
        ["Bachelor's/Master's in IT, business admin or related", "3+ yrs as Functional Analyst", "SAP knowledge is decisive",
         "Production environment experience preferred", "Fluent Dutch and English; French a plus"]),
    ("SAP Enterprise Architect", "Robert Half"): (
        "4-year engagement via Robert Half: define SAP architecture strategy, advise "
        "senior leadership, develop multi-year roadmaps for a Ghent client.",
        ["10-15+ yrs in SAP, complex enterprise environments", "S/4HANA architecture and transformation programs",
         "Enterprise architecture beyond hands-on implementation", "Dutch mandatory; professional English"]),
    ("SAP BDC Lead", "KPMG Belgium"): (
        "Director-level: build/scale KPMG's SAP Business Data Cloud practice, driving "
        "go-to-market and client delivery across BE-NL.",
        ["8+ yrs SAP, data integration or cloud transformation", "Strong SAP BTP and SAP Business Data Cloud knowledge",
         "SAP BW/4HANA, Datasphere, analytics tools", "Plus: Azure Synapse, Snowflake, Databricks",
         "Fluent English; working Dutch and/or French"]),
    ("Senior SAP Sales & Distribution Consultant", "delaware BeLux"): (
        "Design/implement sales and distribution processes for clients, translating "
        "requirements into SAP SD configuration and running trainings.",
        ["5+ yrs SAP SD consulting experience", "Sales/distribution process and best-practice expertise",
         "English and Dutch or French", "Order-to-cash flow knowledge (pricing, ATP, MRP, credit mgmt)"]),
    ("Responsable Fonctionnel SAP Quality / WMS", "Safran"): (
        "SAP Functional Manager for Quality/Warehouse Management at Safran Aero "
        "Boosters (Herstal): S/4HANA implementation plus post-go-live support.",
        ["Master's in IT or equivalent", "8+ yrs SAP with strong QM module expertise; eWM preferred",
         "S/4HANA implementation experience required", "Industrial production sector background",
         "36 hrs/week, 1-2 days remote"]),
    ("Responsable Fonctionnel SAP Finance", "Safran"): (
        "SAP Finance Functional Manager at Safran Aero Boosters (Herstal): S/4HANA "
        "implementation, offshore team management, accounting/controlling processes.",
        ["Master's in IT or equivalent", "10+ yrs with SAP FI-CO",
         "S/4HANA implementation leadership experience", "Belgian GAAP and IFRS knowledge", "Fluent French and English"]),
    ("SAP Architect (SD/MM)", "Amon"): (
        "Strategic bridge between global business objectives and system landscape at "
        "Lotus Bakeries, focused on Sales, Logistics & Purchasing (Kaprijke).",
        ["Deep SAP SD, MM knowledge; integration with FICO, WM, IBP", "Proven SAP end-to-end rollout / project management experience",
         "Pricing, VAT, intercompany flows, rebate management familiarity", "Willingness to travel; hybrid Belgium/South Africa team"]),
    ("SAP FICO Consultant", "Deloitte Belgium"): (
        "Advise clients on SAP finance solutions within Deloitte's Tax Technology "
        "Consulting team, bridging IT and tax professionals.",
        ["3-5 yrs relevant SAP finance experience", "Dutch and English proficiency (other EU languages valued)",
         "Strong stakeholder communication", "Mid-Senior level; full-time or 80-90%"]),
    ("SAP Business One Consultant", "Amaris Consulting"): (
        "Provide hands-on technical/functional guidance to stabilize SAP Business One "
        "systems for a life-sciences client in Herstal.",
        ["Master's in Engineering, Life Sciences, IT, Business or related", "Proven ERP/SAP Business One experience",
         "Supply chain systems + regulated industry (pharma/GMP/GDP) background", "Fluent English and French"]),
    ("SAP Application Security Consultant", "BDO Belgium"): (
        "Lead role design/SoD initiatives for SAP ECC and S/4HANA: SoD rulebooks, user "
        "remediation, access control assessments.",
        ["Proven SAP role design/user assignment experience (ECC, S/4HANA)", "SoD rulebook creation/analysis with risk tuning",
         "IT/process control frameworks knowledge", "ISO 27001, NIST CSF, NIS2, DORA familiarity",
         "1-5 yrs relevant experience; CISSP/CISM/CISA/SAP GRC a plus"]),
    ("SAP Global Trade Services Consultant", "delaware BeLux"): (
        "Implement/manage SAP GTS solutions for international clients: analyze trade "
        "processes, configure, train, provide post-go-live support.",
        ["5+ yrs global trade/customs/logistics/supply chain", "Hands-on SAP GTS (end user, key user or consultant)",
         "Full implementation project experience", "SAP SD integration knowledge preferred", "English plus Dutch or French"]),
    ("SAP Enterprise Architect", "KPMG Belgium"): (
        "Design end-to-end SAP solutions and lead client transformations, expanding "
        "KPMG's growing SAP practice in Zaventem.",
        ["6+ yrs hands-on SAP Integration (BTP Integration Suite, CPI, PI/PO)", "S/4HANA architecture experience/willingness to develop",
         "S/4HANA or BTP certification", "Coaching/mentoring skills", "Willingness to travel internationally"]),
    ("SAP Archiving & ILM Manager", "Deloitte Belgium"): (
        "Lead enterprise data archiving and Information Lifecycle Management across "
        "complex SAP landscapes, focused on compliance and optimization.",
        ["7+ yrs SAP, 3+ yrs specifically Archiving/ILM", "Deep SAP Data Archiving, S/4HANA, ECC knowledge (Jivs, OpenText InfoArchive)",
         "GDPR and regulatory retention expertise", "End-to-end ILM deployment track record", "SAP/ISO 27001 certifications valued"]),
    ("SAP EWM/WM Consultant", "Accenture Belgium"): (
        "Support warehouse management solution design, configuration and delivery: "
        "module customization, client collaboration, training.",
        ["Bachelor's/Master's in IT, Supply Chain, Engineering or related", "Multiple years practical EWM/WM module experience",
         "Fluent English and either French or Dutch (mandatory)", "Demonstrated S/4HANA project involvement"]),
    ("SAP S/4HANA Functional Analyst & Developer", "Talan"): (
        "Support S/4HANA implementations/migrations/enhancements maintaining clean "
        "core architecture, bridging business requirements and technical solutions.",
        ["3+ yrs S/4HANA functional analysis/development, ECC-to-S/4HANA migration", "ABAP Cloud, CDS Views, Fiori/UI5, RAP/CAP, SAP BTP",
         "Module proficiency: FI/CO, MM, SD, PP, IS-U", "RISE with SAP, SAP Activate, clean core understanding",
         "English and French; Dutch advantageous"]),
    ("SAP Developer", "Kingfisher Recruitment"): (
        "Optimize the SAP environment for a major coffee brand in Oud-Turnhout, "
        "preparing for a future S/4HANA migration.",
        ["Bachelor's/Master's in IT", "Previous relevant SAP work experience", "SAP ABAP programming knowledge",
         "Strong analytical thinking", "32 vacation days, remote options"]),
    ("SAP Solution Architect", "Sopra Steria"): (
        "Guide organizations through digital transformation, converting business "
        "challenges into scalable SAP architectures.",
        ["5+ yrs in a similar SAP-focused role", "Deep SAP solution architecture, integration, infrastructure design",
         "Security/performance/reliability requirements experience", "Fluent Dutch and/or French, plus English"]),
    ("SAP PS Senior Manager", "Accenture Belgium"): (
        "Lead the Belgian SAP Project Systems (PS) Practice: business development, "
        "client delivery, team mentorship, RISE with SAP/S/4HANA initiatives.",
        ["5-10 yrs IT/business consulting with SAP projects", "Extensive SAP PS delivery exposure; 1+ yr account/delivery management",
         "Bachelor's minimum (Master's preferred)", "Fluent Dutch and English", "Comfort with matrix structures and client travel"]),
    ("Functional Analyst SAP Logistics (Sr)", "John Cockerill"): (
        "Functional expert supporting operational stability and S/4HANA "
        "transformation across 50+ legal entities; evolving into a referent role.",
        ["8+ yrs as SAP Functional Analyst, logistics focus", "Deep PS, MM, SD, PP, PM, QM module expertise; FI & CO understanding",
         "S/4HANA experience preferred", "End-to-end process vision, multi-module integration", "Ability to work with offshore teams"]),
    ("SAP Supply Chain Planning Specialist", "Deloitte Belgium"): (
        "Join Deloitte's SAP Supply Chain Planning team: manage planning projects "
        "using PP/DS, aATP, IBP, from implementation to training.",
        ["2-4 yrs project experience with IBP or PP/DS implementation", "Strong supply chain planning knowledge",
         "University degree or equivalent experience", "Fluent English (Dutch/French advantageous)", "EU work authorization required"]),
    ("SAP Specialist", "Indaver"): (
        "Support waste-management operations: analyze requirements, implement SAP "
        "across order-to-cash and logistics in a post-S/4HANA-migration environment.",
        ["Bachelor's or equivalent professional experience", "3+ yrs in SAP environments (IT background preferred)",
         "Fluent Dutch and English", "Plus: S/4HANA implementation, ABAP, SAP Query, Fiori"]),
    ("SAP Sales", "Match Profiler"): (
        "SAP Sales/Logistics role for a multinational client via Match Profiler; "
        "hybrid/remote with travel, based in Kaprijke.",
        ["Master's degree", "5+ yrs Sales, Logistics & Purchasing ICT in multinationals", "Strong SAP SD & MM, integration with FICO/WM/IBP",
         "1+ SAP end-to-end rollout participation", "€65,000–€75,000/yr + benefits; relocation support for EU citizens"]),
    ("Lead Consultant – SAP Manufacturing", "Emixa"): (
        "Bridge business and IT, guiding clients through production process "
        "optimization; lead project teams, mentor junior consultants.",
        ["5+ yrs relevant experience with successful project completions", "1+ end-to-end SAP implementation",
         "Deep PP, PP-PI, PP-DS, shopfloor integration, QM expertise", "Fluent Dutch and English; French required",
         "Category B driver's license"]),
    ("SAP Solutions Specialist", "Deloitte Belgium"): (
        "Design/develop/implement enterprise SAP S/4HANA solutions across the full "
        "lifecycle, from requirements through post-go-live support.",
        ["2-8 yrs SAP consulting experience with S/4HANA go-live involvement", "Deep expertise in 1+ module (FI/CO, MM, SD, GTS, EWM, TM, BI/BW, security, HCM)",
         "Fluent English; French or Dutch valued", "Right to work in Belgium required", "Openness to international assignments"]),
    ("SAP Team Lead", "Crop's"): (
        "Lead SAP SD/MM operations while driving the transition toward Clean Core and "
        "BTP for a food & beverage company in West Flanders.",
        ["SAP SD/MM expertise with architectural mindset", "Hands-on technical + analytical strength",
         "Team/project leadership experience (SuccessFactors, EDI)", "Fluent Dutch, English and French"]),
    ("SAP-PP & MES Consultant", "stow Group"): (
        "Design/implement/support production planning modules (ECC and S/4HANA); lead "
        "an MES rollout focused on OEE, automation, quality metrics.",
        ["SAP MM/PP module knowledge with MES system familiarity", "Production execution + cross-functional SAP integration background",
         "Requirements gathering → system configuration translation", "Plus: SD, FICO, IM modules familiarity"]),
    ("Consultant(e) SAP S/4HANA (M/F)", "mc2i"): (
        "Lead digital transformation for enterprise clients in Brussels: design "
        "processes, configure modules, manage data migrations (Greenfield/Brownfield).",
        ["Bachelor's (Engineering/Business/University) + SAP experience or aptitude", "S/4HANA + project methodology (Agile/Scrum/Waterfall) expertise",
         "Fluent English and French; Dutch advantageous", "Interest in GenAI, Cloud, Sustainability SAP solutions"]),
    ("Solution Architect SAP", "Madison Recruitment"): (
        "Design/develop SAP architecture (SAP EWM, S/4HANA) for an international "
        "logistics company's digital transformation, via Madison Recruitment.",
        ["7+ yrs in SAP environments, Solution Architect background preferred", "SAP EWM hands-on expertise essential",
         "Fluent English", "Alternative path: several years as Functional Analyst SAP EWM"]),
    ("SAP PP Consultant", "delaware BeLux"): (
        "Help manufacturers optimize production/planning via SAP S/4HANA: design "
        "workflows, configure modules, coach teams (Flemish Region).",
        ["5+ yrs as SAP PP Consultant", "Strong manufacturing/production planning expertise",
         "Fluent Dutch or French; solid English", "Plus: PPDS, IBP, or MES systems knowledge"]),
    ("SAP S/4HANA Finance Consultant", "Eursap"): (
        "Contract role for an international client (France, remote) via Eursap's SAP-only "
        "feed. The listing itself gives only Job ID/rate/start date — no detailed "
        "requirements text is published; ask Eursap directly for the full spec.",
        []),
    ("SAP BTP Techno-Functional Consultant", "Eursap"): (
        "Contract role for an international client (France, remote) via Eursap's SAP-only "
        "feed. The listing itself gives only Job ID/rate/start date — no detailed "
        "requirements text is published; ask Eursap directly for the full spec.",
        []),
    ("SAP Project Manager / Trainer", "Eursap"): (
        "Permanent role (remote) via Eursap's SAP-only feed, PM/Trainer scope. The "
        "listing itself gives only Job ID/salary/start date — no detailed requirements "
        "text is published; ask Eursap directly for the full spec.",
        []),
}


EURSAP_ROLES = [
    ("SAP S/4HANA Finance Consultant", "Eursap", "France, remote", "Start date 5 Oct 2026 shown — not a posting date"),
    ("SAP BTP Techno-Functional Consultant", "Eursap", "France, remote", "Start date 5 Oct 2026 shown — not a posting date"),
    ("SAP Project Manager / Trainer", "Eursap", "Remote", "Start date 1 Sep 2026 shown — not a posting date"),
]

HEADER = ["Function", "Procurement party / internal job", "Contact person", "Contact details",
          "When posted", "Location", "Job description", "Key requirements"]

HEADER_FILL = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def write_sheet(ws, title, roles):
    ws.title = title
    ws.append(HEADER)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for role_title, company, location, posted in roles:
        party_label, person, details = contact(company)
        description, requirements = JOB_DETAILS.get(
            (role_title, company), ("Not fetched this pass — see the LinkedIn posting directly.", [])
        )
        req_text = "\n".join(f"- {r}" for r in requirements) if requirements else "—"
        ws.append([role_title, party_label, person, details, posted, location, description, req_text])
    widths = [38, 32, 24, 42, 20, 20, 46, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    ws.freeze_panes = "A2"


def build_xlsx():
    wb = Workbook()
    write_sheet(wb.active, "Netherlands", NL_ROLES)
    write_sheet(wb.create_sheet("Belgium"), "Belgium", BE_ROLES)
    write_sheet(wb.create_sheet("Eursap (remote feed)"), "Eursap", EURSAP_ROLES)

    notes = wb.create_sheet("Read me")
    notes.append(["SAP Job Search — Belgium & Netherlands: contact directory notes"])
    notes["A1"].font = Font(bold=True, size=13)
    lines = [
        "",
        f"Generated {REPORT_DATE}. Re-verify contact details before calling — recruiter",
        "phone numbers, named contacts and open roles all change frequently.",
        "",
        "\"Agency\" in column 2 = a genuine third-party procurement/staffing party",
        "(Eursap, Michael Page, Robert Half, Madison Recruitment, Kingfisher Recruitment,",
        "RED Global, Match Profiler, ...) that recruits on behalf of an undisclosed or",
        "named client, and does publish a general desk you can call.",
        "",
        "\"Direct employer\" = the company hiring for its own SAP team directly (Accenture,",
        "Deloitte, Capgemini, KPMG, delaware, Sopra Steria, BCG Platinion, Stedin, and",
        "most of the smaller single-listing companies). None of these publish a named",
        "recruiter or direct phone tied to a specific vacancy — you apply through their",
        "careers portal and a recruiter is assigned to you afterwards. Where a company's",
        "general HR/office switchboard number was published, it is listed; that reaches",
        "reception, not necessarily the SAP recruiting team directly.",
        "",
        "Only one specific, verifiably-named recruiter was found this pass: Tuur",
        "Vandeurzen at Madison Recruitment (Belgium). All other \"Contact person\" cells",
        "read \"—\" rather than a fabricated name.",
        "",
        "Companies not individually researched this pass (single-listing employers with",
        "no contact lookup done): Alliander, Brabant Water, myBrand | Conclusion, Nyrstar,",
        "Panda International, vidaXL, Applied Medical, Sibelga, Aures, Safran, Amon,",
        "Amaris Consulting, BDO Belgium, Talan, John Cockerill, Indaver, Emixa, Crop's,",
        "stow Group, mc2i, JoBBsquare België, ThoughtLabs Belgium. For these, apply via",
        "the company's own careers page.",
    ]
    for line in lines:
        notes.append([line])
    notes.column_dimensions["A"].width = 90
    for row in notes.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(str(XLSX_OUTPUT_PATH))
    print(f"XLSX report saved to: {XLSX_OUTPUT_PATH}")


if __name__ == "__main__":
    build_xlsx()
